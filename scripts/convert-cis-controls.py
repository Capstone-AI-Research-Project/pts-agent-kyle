#!/usr/bin/env python3
"""
Convert CIS Controls v8.1.2 Excel to JSON
Input:  assets/CIS_Controls_Version_8.1.2_March_2025.xlsx
Output: assets/cis-controls-v8.1.2.json

Run: python3 scripts/convert-cis-controls.py
Requires: pip install openpyxl
"""

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

ASSETS_DIR = Path(__file__).parent.parent / "assets"
INPUT_FILE = ASSETS_DIR / "CIS_Controls_Version_8.1.2_March_2025.xlsx"
OUTPUT_FILE = ASSETS_DIR / "cis-controls-v8.1.2.json"

STOPWORDS = {
    'a', 'an', 'the', 'and', 'or', 'of', 'in', 'on', 'at', 'to', 'for',
    'is', 'are', 'was', 'be', 'by', 'with', 'from', 'that', 'this', 'it',
    'all', 'as', 'up', 'out', 'if', 'no', 'not', 'can', 'use', 'per',
}

# Security function name → NIST CSF function ID
SEC_FUNC_TO_CSF = {
    'identify': 'ID',
    'protect': 'PR',
    'detect': 'DE',
    'respond': 'RS',
    'recover': 'RC',
    'governance': 'GV',
    'govern': 'GV',
}

# Standard CIS Controls v8 titles (fallback if not in Excel)
CONTROL_TITLES = {
    1:  "Inventory and Control of Enterprise Assets",
    2:  "Inventory and Control of Software Assets",
    3:  "Data Protection",
    4:  "Secure Configuration of Enterprise Assets and Software",
    5:  "Account Management",
    6:  "Access Control Management",
    7:  "Continuous Vulnerability Management",
    8:  "Audit Log Management",
    9:  "Email and Web Browser Protections",
    10: "Malware Defenses",
    11: "Data Recovery",
    12: "Network Infrastructure Management",
    13: "Network Monitoring and Defense",
    14: "Security Awareness and Skills Training",
    15: "Service Provider Management",
    16: "Application Software Security",
    17: "Incident Response Management",
    18: "Penetration Testing",
}


def extract_keywords(text):
    """Extract meaningful keywords from text."""
    if not text:
        return []
    words = re.findall(r'[a-zA-Z]{3,}', str(text).lower())
    return list(set(w for w in words if w not in STOPWORDS))


def parse_ig_check(value):
    """Return True if the cell value indicates an IG checkbox is marked."""
    if value is None:
        return False
    sv = str(value).strip().lower()
    return sv in ('x', 'yes', 'true', '1', '✓', '✔', 'y', 'ig1', 'ig2', 'ig3')


def find_header_row(sheet):
    """Scan the first 20 rows to find the column header row."""
    for row_idx, row in enumerate(sheet.iter_rows(max_row=20, values_only=True), 1):
        row_str = ' '.join(str(c).lower() for c in row if c is not None)
        if ('safeguard' in row_str or 'asset type' in row_str) and 'title' in row_str:
            return row_idx, list(row)
    return None, None


def col_idx(headers, *candidates):
    """Find first column index whose header contains any candidate substring."""
    lower = [str(h).lower().strip() if h is not None else '' for h in headers]
    for cand in candidates:
        cand_l = cand.lower()
        for i, h in enumerate(lower):
            if cand_l in h:
                return i
    return None


def convert():
    if not INPUT_FILE.exists():
        print(f"ERROR: Input file not found: {INPUT_FILE}")
        sys.exit(1)

    wb = openpyxl.load_workbook(INPUT_FILE)
    print(f"Workbook sheets: {wb.sheetnames}")

    # Choose the most likely sheet
    sheet = None
    for name in wb.sheetnames:
        nl = name.lower()
        if any(kw in nl for kw in ['safeguard', 'control', 'cis']):
            sheet = wb[name]
            print(f"Selected sheet: {name!r}")
            break
    if sheet is None:
        sheet = wb.active
        print(f"Falling back to active sheet: {sheet.title!r}")

    header_idx, headers = find_header_row(sheet)
    if headers is None:
        print("ERROR: Could not find header row. Printing first 10 rows for debugging:")
        for i, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), 1):
            print(f"  Row {i}: {list(row)}")
        sys.exit(1)

    print(f"\nHeader row at row {header_idx}: {headers}")

    # Map column names
    C_SG_ID    = col_idx(headers, 'safeguard', 'sg #', 'sg#', 'safeguard #', 'control #', 'id')
    C_TITLE    = col_idx(headers, 'title', 'safeguard title', 'name')
    C_DESC     = col_idx(headers, 'description', 'desc', 'overview', 'procedures', 'recommendation')
    C_ASSET    = col_idx(headers, 'asset type', 'asset')
    C_SEC_FUNC = col_idx(headers, 'security function', 'sec function', 'function')
    C_IG1      = col_idx(headers, 'ig1', 'ig 1', 'implementation group 1', 'group 1')
    C_IG2      = col_idx(headers, 'ig2', 'ig 2', 'implementation group 2', 'group 2')
    C_IG3      = col_idx(headers, 'ig3', 'ig 3', 'implementation group 3', 'group 3')

    print(f"\nColumn mapping:")
    print(f"  Safeguard ID: col {C_SG_ID}")
    print(f"  Title:        col {C_TITLE}")
    print(f"  Description:  col {C_DESC}")
    print(f"  Asset Type:   col {C_ASSET}")
    print(f"  Security Fn:  col {C_SEC_FUNC}")
    print(f"  IG1: {C_IG1}  IG2: {C_IG2}  IG3: {C_IG3}")

    if C_SG_ID is None or C_TITLE is None:
        print("\nERROR: Could not find required columns (Safeguard ID and Title).")
        print("Please check the header row above and adjust col_idx() candidates.")
        sys.exit(1)

    def cell(row, idx):
        """Safely get cell value as stripped string."""
        if idx is None or idx >= len(row):
            return ''
        v = row[idx]
        return str(v).strip() if v is not None else ''

    controls_map = {}

    for row in sheet.iter_rows(min_row=header_idx + 1, values_only=True):
        if all(c is None for c in row):
            continue

        sg_id_raw = cell(row, C_SG_ID)

        # Safeguard IDs are like "1.1", "8.11", "18.5"
        m = re.match(r'^(\d+)\.(\d+)$', sg_id_raw)
        if not m:
            continue

        ctrl_num    = int(m.group(1))
        sg_title    = cell(row, C_TITLE)
        description = cell(row, C_DESC)
        asset_type  = cell(row, C_ASSET)
        sec_func    = cell(row, C_SEC_FUNC)

        # Implementation Groups
        igs = []
        if C_IG1 is not None and C_IG1 < len(row) and parse_ig_check(row[C_IG1]):
            igs.append("IG1")
        if C_IG2 is not None and C_IG2 < len(row) and parse_ig_check(row[C_IG2]):
            igs.append("IG2")
        if C_IG3 is not None and C_IG3 < len(row) and parse_ig_check(row[C_IG3]):
            igs.append("IG3")
        if not igs:
            igs = ["IG1", "IG2", "IG3"]  # Default if detection failed

        nist_func = SEC_FUNC_TO_CSF.get(sec_func.lower(), '')

        # Build keyword list from title + first 200 chars of description
        keywords = extract_keywords(sg_title) + extract_keywords(description[:200])
        keywords = sorted(set(keywords))[:20]

        safeguard = {
            "id":                   sg_id_raw,
            "title":                sg_title,
            "description":          description[:500],
            "asset_type":           asset_type,
            "security_function":    sec_func,
            "nist_csf_function":    nist_func,
            "implementation_groups": igs,
            "keywords":             keywords,
        }

        if ctrl_num not in controls_map:
            controls_map[ctrl_num] = {
                "id":        ctrl_num,
                "title":     CONTROL_TITLES.get(ctrl_num, f"CIS Control {ctrl_num}"),
                "safeguards": [],
            }

        controls_map[ctrl_num]["safeguards"].append(safeguard)

    controls = [controls_map[k] for k in sorted(controls_map.keys())]
    total_sgs = sum(len(c["safeguards"]) for c in controls)

    output = {
        "version":         "8.1.2",
        "date":            "March 2025",
        "description":     "CIS Critical Security Controls v8.1.2 — safeguard-level security recommendations",
        "control_count":   len(controls),
        "safeguard_count": total_sgs,
        "controls":        controls,
    }

    OUTPUT_FILE.write_text(json.dumps(output, indent=2), encoding='utf-8')
    print(f"\n✓ Converted {len(controls)} controls, {total_sgs} safeguards → {OUTPUT_FILE}")

    # IG distribution
    ig1 = sum(1 for c in controls for s in c["safeguards"] if "IG1" in s["implementation_groups"])
    ig2 = sum(1 for c in controls for s in c["safeguards"] if "IG2" in s["implementation_groups"])
    ig3 = sum(1 for c in controls for s in c["safeguards"] if "IG3" in s["implementation_groups"])
    print(f"  IG1 safeguards (quick wins): {ig1}")
    print(f"  IG2 safeguards:              {ig2}")
    print(f"  IG3 safeguards:              {ig3}")


if __name__ == '__main__':
    convert()
